def invoke_GenerateAndUploadAktliste(Arguments_GenerateAndUploadAktliste):
    import os
    import shutil
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from PIL import ImageFont, ImageDraw, Image
    import pandas as pd
    from datetime import datetime 
    import time        
    from SharePointUploader import upload_file_to_sharepoint
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.worksheet.dimensions import ColumnDimension
    from PIL import ImageFont  
    import textwrap
    import math
    from GetDocumentList import sharepoint_client
    import reportlab
    # ReportLab Imports
    from reportlab.pdfgen import canvas as reportlab_canvas
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors as reportlab_colors
    from reportlab.platypus import Table as ReportTable, TableStyle as ReportTableStyle, Paragraph, SimpleDocTemplate, Frame, PageTemplate
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle




    # henter in_argumenter:
    dt_AktIndex = Arguments_GenerateAndUploadAktliste.get("in_dt_AktIndex")
    Sagsnummer = Arguments_GenerateAndUploadAktliste.get("in_Sagsnummer")
    DokumentlisteDatoString = Arguments_GenerateAndUploadAktliste.get("in_DokumentlisteDatoString")
    RobotUserName = Arguments_GenerateAndUploadAktliste.get("in_RobotUserName")
    RobotPassword = Arguments_GenerateAndUploadAktliste.get("in_RobotPassword")
    Sagstitel = Arguments_GenerateAndUploadAktliste.get("in_SagsTitel")
    SharePointAppID = Arguments_GenerateAndUploadAktliste.get("in_SharePointAppID")
    SharePointTenant = Arguments_GenerateAndUploadAktliste.get("in_SharePointTenant")
    SharePointURL = Arguments_GenerateAndUploadAktliste.get("in_SharePointURL")
    Overmappe = Arguments_GenerateAndUploadAktliste.get("in_Overmappe")
    Undermappe = Arguments_GenerateAndUploadAktliste.get("in_Undermappe")
    GoUsername = Arguments_GenerateAndUploadAktliste.get("in_GoUsername")
    GoPassword = Arguments_GenerateAndUploadAktliste.get("in_GoPassword")
    tenant = Arguments_GenerateAndUploadAktliste.get("tenant")
    client_id = Arguments_GenerateAndUploadAktliste.get("client_id")
    thumbprint = Arguments_GenerateAndUploadAktliste.get("thumbprint")
    cert_path = Arguments_GenerateAndUploadAktliste.get("cert_parth")
    
    ctx = sharepoint_client(tenant, client_id, thumbprint, cert_path, SharePointURL)
    

    def create_excel(data_table, file_path):
        try:
            # Convert UiPath DataTable (dt_AktIndex) to Pandas DataFrame
            data = pd.DataFrame(data_table)

            # Check if the DataFrame is empty
            if data.empty:
                raise ValueError("The provided DataTable is empty.")

            # Replace all NaN, pd.NA, and similar with None to make them empty in Excel
            data = data.where(pd.notnull(data), None)
            data = data.replace({pd.NA: None, "nan": None, "NaN": None, "None": None})

            # Format all dates in the DataFrame to dd-MM-yyyy
            for col in data.columns:
                if pd.api.types.is_datetime64_any_dtype(data[col]):
                    data[col] = data[col].dt.strftime("%d-%m-%Y")

            # Create workbook and worksheet
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Aktliste"

            # Insert headers
            worksheet.row_dimensions[1].height = 20  # Adjust header row height
            num_rows, num_cols = data.shape
            for col_index, column_name in enumerate(data.columns):
                col_letter = get_column_letter(col_index + 1)
                cell = worksheet.cell(row=1, column=col_index + 1, value=column_name)
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

                # Set column width dynamically and specifically for "Filnavn"
                if column_name == "Filnavn":
                    worksheet.column_dimensions[col_letter].width = 79  # Set "Filnavn" column to 79
                else:
                    worksheet.column_dimensions[col_letter].width = max(15, len(column_name) + 5)

            # Insert data rows with wrap text and dynamic row heights
            for row_index, row in data.iterrows():
                max_lines = 1  # Track maximum lines per row to adjust row height
                for col_index, value in enumerate(row):
                    cell = worksheet.cell(row=row_index + 2, column=col_index + 1, value=value)
                    cell.alignment = Alignment(wrap_text=True, vertical="center")

                    # Calculate line count if value exists
                    if value and col_index == 1:  # Specific to "Filnavn" column
                        lines = textwrap.wrap(str(value), width=70)  # Adjust line wrap width
                        max_lines = max(max_lines, len(lines))
                    if col_index == 4:  # Adjust height for DOKID
                        worksheet.row_dimensions[row_index + 2].height = 20
                # Dynamically adjust row height based on line count
                worksheet.row_dimensions[row_index + 2].height = 15 * max_lines

            # Define table range
            table_range = f"A1:{get_column_letter(num_cols)}{num_rows + 1}"
            table = Table(displayName="AktTable", ref=table_range)

            # Apply blue and white theme using built-in styles
            style = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            table.tableStyleInfo = style
            worksheet.add_table(table)

            # Save workbook
            workbook.save(file_path)

            # Move file to Downloads folder
            downloads_folder = os.path.join("C:\\Users", os.getlogin(), "Downloads")
            if not os.path.exists(downloads_folder):
                os.makedirs(downloads_folder)
            destination_path = os.path.join(downloads_folder, os.path.basename(file_path))
            shutil.move(file_path, destination_path)
            print(f"File moved to {destination_path}")
        
        except Exception as e:
            raise Exception(f"Error creating Excel file: {str(e)}")

    # Example usage
    AktlisteFileName = f"Aktliste - {Sagsnummer} - {DokumentlisteDatoString}.xlsx"
    file_path = AktlisteFileName
    create_excel(dt_AktIndex, file_path)

        
    downloads_path = os.path.join("C:\\Users", os.getlogin(), "Downloads")
    file_path = os.path.join(downloads_path, AktlisteFileName)


    upload_file_to_sharepoint(
        site_url=SharePointURL,
        Overmappe=Overmappe,
        Undermappe=Undermappe,
        file_path=file_path,
        ctx
    )


    def wrap_text(text, max_chars):
        if pd.isna(text): 
            return ""
        if not isinstance(text, str):
            text = str(text)
        words = text.split()
        wrapped_lines = []
        line = ""
        for word in words:
            if len(line) + len(word) + 1 <= max_chars:
                line += " " + word if line else word
            else:
                wrapped_lines.append(line)
                line = word
        if line:
            wrapped_lines.append(line)
        return "<br/>".join(wrapped_lines)

    def excel_to_pdf(excel_path, image_path, output_pdf_path, sags_id, my_date_string):
        try:
            df = pd.read_excel(excel_path)
            
            # PDF Setup
            page_width, page_height = landscape(A4)
            margin = 40


            # Define styles
            styles = getSampleStyleSheet()

            header_style = ParagraphStyle(
                'header_style',
                parent=styles['Normal'],
                fontName='Helvetica-Bold',
                fontSize=10,
                textColor=reportlab_colors.white,
                alignment=1,  # CENTER
                leading=12,
                spaceAfter=5,
            )

            cell_style = ParagraphStyle(
                'cell_style',
                parent=styles['Normal'],
                fontName='Helvetica',
                fontSize=8,
                textColor=reportlab_colors.black,
                alignment=1,  # CENTER
                leading=10,
                spaceAfter=2,
            )

            # Column configuration
            column_widths = [50, 150, 80, 70, 75, 55, 50, 65, 70, 100]
            char_limits = [10, 30, 15, 12, 15, 10, 9, 12, 12, 20]

            headers = ["Akt ID", "Filnavn", "Kategori", "Dato", "Dok ID", "Bilag til Dok ID", 
                    "Bilag", "Omfattet af aktindsigt", "Gives der aktindsigt?", "Begrundelse"]

            # Create header row
            table_data = [[Paragraph(header, header_style) for header in headers]]

            # Add data rows
            for _, row in df.iterrows():
                table_row = [
                    Paragraph(wrap_text(row.get("Akt ID", ""), char_limits[0]), cell_style),
                    Paragraph(wrap_text(row.get("Filnavn", ""), char_limits[1]), cell_style),
                    Paragraph(wrap_text(row.get("Dokumentkategori", ""), char_limits[2]), cell_style),
                    Paragraph(wrap_text(row.get("Dokumentdato", "").strftime("%d-%m-%Y") if isinstance(row.get("Dokumentdato"), pd.Timestamp) else row.get("Dokumentdato", ""), char_limits[3]), cell_style),
                    Paragraph(wrap_text(row.get("Dok ID", ""), char_limits[4]), cell_style),
                    Paragraph(wrap_text(row.get("Bilag til Dok ID", ""), char_limits[5]), cell_style),
                    Paragraph(wrap_text(row.get("Bilag", ""), char_limits[6]), cell_style),
                    Paragraph(wrap_text(row.get("Omfattet af aktindsigt?", ""), char_limits[7]), cell_style),
                    Paragraph(wrap_text(row.get("Gives der aktindsigt?", ""), char_limits[8]), cell_style),
                    Paragraph(wrap_text(row.get("Begrundelse hvis Nej/Delvis", ""), char_limits[9]), cell_style)
                ]
                table_data.append(table_row)

            # Create table
            report_table = ReportTable(table_data, colWidths=column_widths)
            report_table.setStyle(ReportTableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), reportlab_colors.HexColor("#3661D8")),
                ('GRID', (0, 0), (-1, -1), 1, reportlab_colors.black),
                ('BOX', (0, 0), (-1, -1), 1, reportlab_colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ]))

            # Define first page layout
            def first_page(canvas, doc):
                canvas.saveState()
                
                # Draw Image
                image_width = 100
                image_height = 45
                image_x = margin
                image_y = page_height - margin - image_height
                canvas.drawImage(image_path, image_x, image_y, width=image_width, height=image_height)

                # Add Title
                title = f"Aktliste - {sags_id}"
                canvas.setFont("Helvetica-Bold", 14)
                title_y = image_y - 20  # Title below the image
                canvas.drawString(margin, title_y, title)
                
                # Measure the width of the title
                title_width = canvas.stringWidth(title, "Helvetica-Bold", 14)


                    # Add Black Line **RIGHT BELOW** the Title
                line_y = title_y - 5  # 5 units below the title
                canvas.setStrokeColor(reportlab_colors.black)
                canvas.setLineWidth(1)
                canvas.line(margin, line_y, margin + title_width, line_y) 
                
                # Add Date
                date_string = f"Dato for aktindsigt: {my_date_string}"
                canvas.setFont("Helvetica", 10)
                text_width = canvas.stringWidth(date_string, "Helvetica", 10)
                canvas.drawString(page_width - margin - text_width, image_y, date_string)

                canvas.restoreState()

            # Define subsequent pages (only table continues)
            def later_pages(canvas, doc):
                canvas.saveState()
                canvas.restoreState()

            # PDF document setup
            doc = SimpleDocTemplate(output_pdf_path, pagesize=landscape(A4),
                                    leftMargin=margin, rightMargin=margin,
                                    topMargin=margin, bottomMargin=margin)

            # Reserve space at the top for image, title, and date
            table_start_y = page_height - margin - 100  # Adjusted Y position to avoid overlap

            # Define frames (where table goes)
            frame_first_page = Frame(margin, margin, page_width - 2 * margin, table_start_y - margin, id='first_page_table_frame')
            frame_later_pages = Frame(margin, margin, page_width - 2 * margin, page_height - 2 * margin, id='later_page_table_frame')

            # Define page templates
            first_page_template = PageTemplate(id='FirstPage', frames=frame_first_page, onPage=first_page)
            later_page_template = PageTemplate(id='LaterPages', frames=frame_later_pages, onPage=later_pages)
            
            doc.addPageTemplates([first_page_template, later_page_template])

            # Build the PDF with the table content
            doc.build([report_table])

            print(f"PDF saved to {output_pdf_path}")
        
        except Exception as e:
            raise Exception(f"Error in generating pdf: {str(e)}")

    PDFAktlisteFilnavn = f"Aktliste - {Sagsnummer} - {DokumentlisteDatoString}.pdf"

    excel_to_pdf(
            os.path.join(downloads_path, AktlisteFileName),
            os.path.join(os.getcwd(), "aak.jpg"),
            os.path.join(downloads_path, PDFAktlisteFilnavn),
            Sagsnummer,
            DokumentlisteDatoString
        )

    downloads_path = os.path.join("C:\\Users", os.getlogin(), "Downloads")
    pdf_path = os.path.join(downloads_path, PDFAktlisteFilnavn)

    # Upload Excel to Sharepoint
    upload_file_to_sharepoint(
        site_url=SharePointURL,
        Overmappe=Overmappe,
        Undermappe=Undermappe,
        file_path=pdf_path,
        ctx
    )

    #Deleting local files: 
    try:
        os.remove(file_path)
        os.remove(pdf_path)
    except Exception as e:
        raise Exception(f"Error deleting local files: {str(e)}")

    return {
    "out_Text": f"Generated Aktliste",
    }
